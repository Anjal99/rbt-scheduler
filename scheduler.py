#!/usr/bin/env python3.11
"""
RBT Therapy Scheduler Engine
Reads therapist/client data from Excel, generates a valid weekly schedule,
and writes it to the "Current Assignment" sheet.

Usage: python3.11 scheduler.py
"""

import openpyxl
import re
import os
from dataclasses import dataclass
from datetime import time
from typing import Optional


# ── Constants ────────────────────────────────────────────────────────────────

DAYS_ORDER = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
WEEKDAYS = DAYS_ORDER[:5]
DAYS_SET = set(DAYS_ORDER)

SOFT_CAP = 30.0
HARD_CAP = 35.0
FORTY_CAP = 40.0
BREAK_MINUTES = 30
MAX_CHAIN_HOURS = 4.0
HIGH_INTENSITY_MAX = 3.0
LOW_INTENSITY_MAX = 4.0


# ── Data Structures ──────────────────────────────────────────────────────────

@dataclass
class TimeBlock:
    start: time
    end: time

    def duration_minutes(self) -> int:
        return (self.end.hour * 60 + self.end.minute) - (self.start.hour * 60 + self.start.minute)

    def duration_hours(self) -> float:
        return self.duration_minutes() / 60.0

    def overlaps(self, other: 'TimeBlock') -> bool:
        return self.start < other.end and other.start < self.end

    def intersection(self, other: 'TimeBlock') -> Optional['TimeBlock']:
        if not self.overlaps(other):
            return None
        s = max(self.start, other.start)
        e = min(self.end, other.end)
        if s >= e:
            return None
        return TimeBlock(s, e)

    def contains_time(self, t: time) -> bool:
        return self.start <= t < self.end

    def __repr__(self):
        return f"{fmt_time(self.start)}-{fmt_time(self.end)}"


@dataclass
class Therapist:
    name: str
    days: list
    availability: dict          # day -> list[TimeBlock]
    in_home: bool
    preferred_max_hours: Optional[float]
    forty_hour_eligible: bool
    is_float: bool = False
    direct_target: Optional[float] = None
    direct_max: Optional[float] = None
    notes: str = ""
    flexible_days: Optional[int] = None


@dataclass
class Client:
    name: str
    schedule: dict              # day -> TimeBlock
    days: list
    location: str               # "Clinic", "Home", "Hybrid"
    location_by_day: dict       # day -> "Clinic" or "Home"
    intensity: str              # "High" or "Low"
    travel_notes: str = ""
    notes: str = ""


@dataclass
class Assignment:
    client: str
    therapist: str
    day: str
    start: time
    end: time
    location: str
    assignment_type: str        # "Recurring" or "Float"
    notes: str = ""

    def duration_minutes(self) -> int:
        return (self.end.hour * 60 + self.end.minute) - (self.start.hour * 60 + self.start.minute)

    def duration_hours(self) -> float:
        return self.duration_minutes() / 60.0


# ── Utility Functions ────────────────────────────────────────────────────────

def fmt_time(t: time) -> str:
    """Format time as '9:00am', '3:30pm'."""
    h = t.hour
    m = t.minute
    ampm = 'am' if h < 12 else 'pm'
    h12 = h if h <= 12 else h - 12
    if h12 == 0:
        h12 = 12
    if m == 0:
        return f"{h12}{ampm}"
    return f"{h12}:{m:02d}{ampm}"


def time_to_minutes(t: time) -> int:
    return t.hour * 60 + t.minute


def minutes_to_time(m: int) -> time:
    m = max(0, min(m, 23 * 60 + 59))
    return time(m // 60, m % 60)


def time_add_minutes(t: time, mins: int) -> time:
    return minutes_to_time(time_to_minutes(t) + mins)


# ── Parsing ──────────────────────────────────────────────────────────────────

def parse_time(s: str) -> time:
    """Parse '9am', '3:30pm', '12pm', '2:00pm', etc."""
    s = s.strip().lower().replace('.', '')
    m = re.match(r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)', s)
    if not m:
        raise ValueError(f"Cannot parse time: '{s}'")
    hour = int(m.group(1))
    minute = int(m.group(2)) if m.group(2) else 0
    ampm = m.group(3)
    if ampm == 'pm' and hour != 12:
        hour += 12
    elif ampm == 'am' and hour == 12:
        hour = 0
    return time(hour, minute)


def normalize_day(d: str) -> str:
    """Normalize day name to 3-letter form."""
    d = d.strip().rstrip(',').rstrip('.').capitalize()
    mapping = {
        'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed',
        'Thursday': 'Thu', 'Friday': 'Fri', 'Saturday': 'Sat', 'Sunday': 'Sun',
        'Mon': 'Mon', 'Tue': 'Tue', 'Wed': 'Wed', 'Thu': 'Thu',
        'Fri': 'Fri', 'Sat': 'Sat', 'Sun': 'Sun',
        'Tues': 'Tue', 'Weds': 'Wed', 'Thurs': 'Thu', 'Thur': 'Thu',
    }
    return mapping.get(d, d)


def expand_day_range(start_day: str, end_day: str) -> list:
    s = normalize_day(start_day)
    e = normalize_day(end_day)
    if s not in DAYS_SET or e not in DAYS_SET:
        return []
    si = DAYS_ORDER.index(s)
    ei = DAYS_ORDER.index(e)
    return DAYS_ORDER[si:ei + 1]


def parse_days_string(s: str) -> tuple:
    """Parse days string. Returns (list_of_days, flexible_count_or_None)."""
    if not s or str(s).strip() == '' or str(s).strip().lower() == 'none':
        return (list(WEEKDAYS), None)

    s = str(s).strip()

    # "4 days (flexible)"
    flex_m = re.match(r'(\d+)\s*days?\s*\(?\s*flexible\s*\)?', s, re.IGNORECASE)
    if flex_m:
        return (list(WEEKDAYS), int(flex_m.group(1)))

    # Replace slashes with commas
    s = s.replace('/', ', ')

    # Range: "Mon - Fri", "Mon-Fri", "Mon – Fri", "Mon- Fri"
    range_m = re.match(r'^(\w+)\s*[-–]\s*(\w+)\s*$', s.strip())
    if range_m:
        days = expand_day_range(range_m.group(1), range_m.group(2))
        if days:
            return (days, None)

    # Comma-separated
    if ',' in s:
        parts = [normalize_day(p.strip()) for p in s.split(',') if p.strip()]
        valid = [p for p in parts if p in DAYS_SET]
        if valid:
            return (valid, None)

    # Space-separated (but filter out non-day words)
    parts = s.split()
    days = [normalize_day(p) for p in parts if normalize_day(p) in DAYS_SET]
    if days:
        return (days, None)

    return (list(WEEKDAYS), None)


def _extract_days_and_time(seg: str, default_days: list) -> tuple:
    """
    From a segment like "Mon - Fri 3pm - 8pm" or "Mon Wed Fri 12pm - 3pm"
    or just "8am - 5pm", extract (days, time_range_string).
    """
    time_pattern = r'(\d{1,2}(?::\d{2})?\s*(?:am|pm)\s*[-–]\s*\d{1,2}(?::\d{2})?\s*(?:am|pm))'
    m = re.search(time_pattern, seg, re.IGNORECASE)
    if not m:
        return (default_days, None)

    time_part = m.group(1)
    prefix = seg[:m.start()].strip()

    if not prefix:
        return (default_days, time_part)

    # Parse prefix as days — could be "Mon - Fri", "Mon Wed Fri", "Mon", "Tue Thu"
    # First check for range pattern in prefix
    range_m = re.match(r'^(\w+)\s*[-–]\s*(\w+)$', prefix.strip())
    if range_m:
        days = expand_day_range(range_m.group(1), range_m.group(2))
        if days:
            return (days, time_part)

    # Try as space/comma separated days
    day_tokens = re.findall(r'[A-Za-z]+', prefix)
    days = [normalize_day(t) for t in day_tokens if normalize_day(t) in DAYS_SET]
    if days:
        return (days, time_part)

    return (default_days, time_part)


def _parse_time_range(s: str) -> Optional[TimeBlock]:
    """Parse '8am - 5pm' into a TimeBlock."""
    s = s.strip()
    parts = re.split(r'\s*[-–]\s*', s)
    if len(parts) != 2:
        return None
    try:
        start = parse_time(parts[0])
        end = parse_time(parts[1])
        # Detect AM/PM typo: start > end
        if time_to_minutes(start) >= time_to_minutes(end):
            if start.hour >= 12:
                start = time(start.hour - 12, start.minute)
                print(f"  [AUTO-FIX] Corrected likely PM→AM typo: {parts[0]} → {fmt_time(start)}")
        if time_to_minutes(start) < time_to_minutes(end):
            return TimeBlock(start, end)
    except ValueError:
        pass
    return None


def parse_hours_string(hours_str: str, available_days: list) -> dict:
    """Parse therapist hours into {day: [TimeBlock, ...]}."""
    result = {d: [] for d in available_days}

    if not hours_str or str(hours_str).strip() == '' or str(hours_str).strip().lower() == 'none':
        default = TimeBlock(time(8, 0), time(20, 0))
        for d in available_days:
            result[d] = [default]
        return result

    hours_str = str(hours_str).strip()
    segments = re.split(r',\s*', hours_str)

    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        days_for_seg, time_part = _extract_days_and_time(seg, available_days)
        if time_part:
            block = _parse_time_range(time_part)
            if block:
                for d in days_for_seg:
                    if d in result:
                        result[d].append(block)

    return result


def parse_client_schedule(schedule_str: str, days: list) -> dict:
    """Parse client schedule string into {day: TimeBlock}."""
    result = {}
    if not schedule_str or str(schedule_str).strip() == '' or str(schedule_str).strip().lower() == 'none':
        return result

    schedule_str = str(schedule_str).strip()
    segments = re.split(r',\s*', schedule_str)

    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        days_for_seg, time_part = _extract_days_and_time(seg, days)
        if time_part:
            block = _parse_time_range(time_part)
            if block:
                for d in days_for_seg:
                    result[d] = block

    return result


def parse_hybrid_notes(notes: str) -> dict:
    """Parse 'Hybrid: Mon, Wed, Fri In-Home; Tue, Thu Clinic' → {day: location}."""
    result = {}
    if not notes:
        return result
    m = re.search(r'[Hh]ybrid:?\s*(.*)', notes)
    if not m:
        return result

    for part in m.group(1).split(';'):
        part = part.strip()
        if not part:
            continue
        loc = "Home" if re.search(r'in[-\s]?home', part, re.IGNORECASE) else "Clinic"
        for day_tok in re.findall(r'[A-Za-z]+', re.sub(r'(in[-\s]?home|clinic)', '', part, flags=re.IGNORECASE)):
            nd = normalize_day(day_tok)
            if nd in DAYS_SET:
                result[nd] = loc
    return result


def parse_float_notes(notes: str) -> tuple:
    """Returns (is_float, direct_target, direct_max)."""
    if not notes:
        return (False, None, None)
    if 'float' not in notes.lower() and 'lead' not in notes.lower():
        return (False, None, None)
    target = None
    max_h = None
    m = re.search(r'Direct\s*Target:\s*(\d+)', notes, re.IGNORECASE)
    if m:
        target = int(m.group(1))
    m = re.search(r'Direct\s*Max:\s*(\d+)', notes, re.IGNORECASE)
    if m:
        max_h = int(m.group(1))
    return (True, target, max_h)


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_therapists(ws) -> list:
    therapists = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        name = row[0].value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()

        days_str = str(row[1].value) if row[1].value else ''
        hours_str = str(row[2].value) if row[2].value else ''
        in_home_val = str(row[3].value).strip().lower() if row[3].value else 'no'
        in_home = in_home_val in ('yes', 'true', '1')

        pref_max = None
        if row[4].value is not None:
            try:
                pref_max = float(row[4].value)
            except (ValueError, TypeError):
                pass

        forty_str = str(row[5].value).strip().lower() if row[5].value else 'no'
        forty_eligible = forty_str in ('yes', 'true', '1')

        notes = str(row[6].value).strip() if row[6].value else ''

        days, flex = parse_days_string(days_str)
        availability = parse_hours_string(hours_str, days)

        # Skip therapists with no real availability (no days/hours specified)
        has_avail = any(len(blocks) > 0 for blocks in availability.values())
        if not has_avail and (not days_str or days_str.strip().lower() == 'none'):
            print(f"  [SKIP] {name} — no availability specified")
            continue

        is_float, dt, dm = parse_float_notes(notes)

        therapists.append(Therapist(
            name=name, days=days, availability=availability,
            in_home=in_home, preferred_max_hours=pref_max,
            forty_hour_eligible=forty_eligible, is_float=is_float,
            direct_target=dt, direct_max=dm, notes=notes,
            flexible_days=flex
        ))
    return therapists


def load_clients(ws) -> list:
    clients = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        name = row[0].value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()

        schedule_str = str(row[1].value).strip() if row[1].value else ''
        days_str = str(row[2].value).strip() if row[2].value else ''
        location_val = str(row[3].value).strip() if row[3].value else 'Clinic'
        travel_notes = str(row[4].value).strip() if row[4].value else ''
        intensity = str(row[5].value).strip().capitalize() if row[5].value else 'Low'
        notes = str(row[6].value).strip() if row[6].value else ''

        days, _ = parse_days_string(days_str)
        schedule = parse_client_schedule(schedule_str, days)

        # Normalize location
        if 'home' in location_val.lower() and 'hybrid' not in location_val.lower():
            location = 'Home'
        elif 'hybrid' in location_val.lower():
            location = 'Hybrid'
        else:
            location = 'Clinic'

        # Build per-day location map
        if location == 'Hybrid':
            location_by_day = parse_hybrid_notes(notes)
            # Fill in missing days with Clinic default
            for d in days:
                if d not in location_by_day:
                    location_by_day[d] = 'Clinic'
        elif location == 'Home':
            location_by_day = {d: 'Home' for d in days}
        else:
            location_by_day = {d: 'Clinic' for d in days}

        clients.append(Client(
            name=name, schedule=schedule, days=days,
            location=location, location_by_day=location_by_day,
            intensity=intensity, travel_notes=travel_notes, notes=notes
        ))
    return clients


# ── Timeline Builder ─────────────────────────────────────────────────────────

def build_therapist_timelines(therapists: list) -> dict:
    """
    Build {therapist_name: {day: [TimeBlock, ...]}}.
    Uses raw availability — chain limits are enforced dynamically during scheduling.
    """
    timelines = {}
    for t in therapists:
        day_blocks = {}
        for d in t.days:
            day_blocks[d] = list(t.availability.get(d, []))
        timelines[t.name] = day_blocks
    return timelines


# ── Scheduling Helpers ───────────────────────────────────────────────────────

def therapist_weekly_hours(name: str, assignments: list) -> float:
    return sum(a.duration_hours() for a in assignments if a.therapist == name)


def therapist_day_assignments(name: str, day: str, assignments: list) -> list:
    result = [a for a in assignments if a.therapist == name and a.day == day]
    result.sort(key=lambda a: a.start)
    return result


def find_free_slots(therapist_name: str, day: str, timelines: dict,
                    assignments: list) -> list:
    """Find remaining free TimeBlocks for a therapist on a day."""
    blocks = timelines.get(therapist_name, {}).get(day, [])
    if not blocks:
        return []

    occupied = therapist_day_assignments(therapist_name, day, assignments)
    free = []

    for block in blocks:
        remaining = [TimeBlock(block.start, block.end)]
        for a in occupied:
            ablock = TimeBlock(a.start, a.end)
            new_remaining = []
            for r in remaining:
                if not r.overlaps(ablock):
                    new_remaining.append(r)
                else:
                    if r.start < ablock.start:
                        new_remaining.append(TimeBlock(r.start, ablock.start))
                    if r.end > ablock.end:
                        new_remaining.append(TimeBlock(ablock.end, r.end))
            remaining = new_remaining
        free.extend(r for r in remaining if r.duration_minutes() >= 15)

    return free


def chain_span_if_inserted(name: str, day: str, new_start: time, new_end: time,
                           assignments: list) -> float:
    """
    Calculate the total chain span (hours) if a new assignment [new_start, new_end]
    were inserted for this therapist on this day.
    Looks both backward AND forward from the proposed assignment.
    A chain is broken only by a gap >= 30 minutes.
    """
    day_a = therapist_day_assignments(name, day, assignments)

    new_s = time_to_minutes(new_start)
    new_e = time_to_minutes(new_end)

    # Start with the proposed assignment span
    chain_earliest = new_s
    chain_latest = new_e

    # Expand backward: find assignments that chain into new_start
    changed = True
    while changed:
        changed = False
        for a in day_a:
            a_s = time_to_minutes(a.start)
            a_e = time_to_minutes(a.end)
            # Assignment ends near or at chain_earliest (gap < 30 min)
            gap = chain_earliest - a_e
            if -1 <= gap < BREAK_MINUTES and a_s < chain_earliest:
                chain_earliest = a_s
                changed = True

    # Expand forward: find assignments that chain from new_end
    changed = True
    while changed:
        changed = False
        for a in day_a:
            a_s = time_to_minutes(a.start)
            a_e = time_to_minutes(a.end)
            # Assignment starts near or at chain_latest (gap < 30 min)
            gap = a_s - chain_latest
            if -1 <= gap < BREAK_MINUTES and a_e > chain_latest:
                chain_latest = a_e
                changed = True

    return (chain_latest - chain_earliest) / 60.0


def travel_buffer_ok(name: str, day: str, new_start: time, new_end: time,
                     new_loc: str, assignments: list) -> bool:
    """Check 30-min travel buffer for in-home sessions."""
    if new_loc != 'Home':
        return True

    for a in therapist_day_assignments(name, day, assignments):
        if a.location != 'Home':
            continue
        # Check gap between sessions
        gap_after = time_to_minutes(new_start) - time_to_minutes(a.end)
        gap_before = time_to_minutes(a.start) - time_to_minutes(new_end)
        # If sessions are adjacent or overlapping, that's the same visit — OK
        # If there's a small gap (0 < gap < 30), that's a travel violation
        if 0 < gap_after < BREAK_MINUTES:
            return False
        if 0 < gap_before < BREAK_MINUTES:
            return False
    return True


def score_therapist(t: Therapist, client: Client, day: str,
                    overlap: TimeBlock, assignments: list) -> float:
    """Score a therapist candidate. Higher = better match."""
    score = 0.0
    weekly = therapist_weekly_hours(t.name, assignments)

    # Strong preference for continuity (same therapist for same client)
    has_client = any(a.therapist == t.name and a.client == client.name for a in assignments)
    if has_client:
        score += 100

    # Prefer same therapist on same day for same client (avoids unnecessary switches)
    has_client_today = any(
        a.therapist == t.name and a.client == client.name and a.day == day
        for a in assignments
    )
    if has_client_today:
        score += 50

    # Prefer larger overlap (more coverage in one shot)
    score += overlap.duration_hours() * 20

    # Prefer lower workload (distribute evenly)
    score -= weekly * 1.0

    # Penalize float/lead therapists (save for gaps)
    if t.is_float:
        score -= 60

    # Penalize therapists over their preferred max
    if t.preferred_max_hours and weekly >= t.preferred_max_hours:
        score -= 40

    # Penalize near hard cap
    if weekly >= HARD_CAP:
        score -= 200
    elif weekly >= SOFT_CAP:
        score -= 20

    return score


# ── Main Scheduling Algorithm ────────────────────────────────────────────────

def try_assign_slot(client: Client, day: str, remaining_start: time,
                    remaining_end: time, therapists: list, timelines: dict,
                    assignments: list, relaxed: bool = False) -> Optional[Assignment]:
    """
    Try to find the best therapist for a client time slot.
    If relaxed=True, allows therapists up to FORTY_CAP even without preferred max check.
    """
    remaining_block = TimeBlock(remaining_start, remaining_end)
    day_loc = client.location_by_day.get(day, 'Clinic')
    intensity_max = HIGH_INTENSITY_MAX if client.intensity == 'High' else LOW_INTENSITY_MAX

    best_t = None
    best_overlap = None
    best_score = -9999

    for t in therapists:
        if day not in t.days:
            continue
        if day_loc == 'Home' and not t.in_home:
            continue

        weekly = therapist_weekly_hours(t.name, assignments)
        # Workload filters (relaxed mode allows up to FORTY_CAP for all eligible)
        if t.is_float and t.direct_max and weekly >= t.direct_max:
            continue
        if not relaxed:
            if not t.forty_hour_eligible and weekly >= HARD_CAP:
                continue
            if weekly >= FORTY_CAP:
                continue
        else:
            if weekly >= FORTY_CAP:
                continue

        free_slots = find_free_slots(t.name, day, timelines, assignments)

        for slot in free_slots:
            overlap = remaining_block.intersection(slot)
            if not overlap or overlap.duration_minutes() < 15:
                continue

            # Cap duration by intensity limit first
            max_mins = int(intensity_max * 60)
            capped_mins = min(overlap.duration_minutes(), max_mins)
            capped_end = time_add_minutes(overlap.start, capped_mins)
            if time_to_minutes(capped_end) > time_to_minutes(overlap.end):
                capped_end = overlap.end

            # Binary search for max duration that doesn't violate chain limit
            # Start with the intensity-capped duration and shrink if needed
            while capped_mins >= 15:
                test_end = time_add_minutes(overlap.start, capped_mins)
                if time_to_minutes(test_end) > time_to_minutes(overlap.end):
                    test_end = overlap.end
                chain_span = chain_span_if_inserted(
                    t.name, day, overlap.start, test_end, assignments
                )
                if chain_span <= MAX_CHAIN_HOURS + 0.01:
                    break
                capped_mins -= 15  # Shrink by 15-min increments
            else:
                continue  # Can't fit even 15 min without chain violation

            capped_end = time_add_minutes(overlap.start, capped_mins)
            if time_to_minutes(capped_end) > time_to_minutes(overlap.end):
                capped_end = overlap.end
            capped = TimeBlock(overlap.start, capped_end)
            if capped.duration_minutes() < 15:
                continue

            # Travel buffer check
            if not travel_buffer_ok(t.name, day, capped.start, capped.end,
                                    day_loc, assignments):
                continue

            s = score_therapist(t, client, day, capped, assignments)
            if s > best_score:
                best_score = s
                best_t = t
                best_overlap = capped

    if best_t and best_overlap:
        # Determine assignment type
        atype = "Recurring"
        if best_t.is_float:
            rec_hours = sum(
                a.duration_hours() for a in assignments
                if a.therapist == best_t.name and a.assignment_type == "Recurring"
            )
            if best_t.direct_target and rec_hours >= best_t.direct_target:
                atype = "Float"

        return Assignment(
            client=client.name,
            therapist=best_t.name,
            day=day,
            start=best_overlap.start,
            end=best_overlap.end,
            location=day_loc,
            assignment_type=atype,
            notes=""
        )
    return None


def schedule_all(therapists: list, clients: list, timelines: dict) -> tuple:
    """
    Greedy, client-first scheduling with a gap-filling second pass.
    Returns (assignments, warnings).
    """
    assignments = []
    warnings = []

    # Sort clients: hardest to place first
    def difficulty(c: Client) -> tuple:
        total_hours = sum(tb.duration_hours() for tb in c.schedule.values())
        is_home = 1 if c.location in ('Home', 'Hybrid') else 0
        is_high = 1 if c.intensity == 'High' else 0
        return (-is_high, -is_home, -total_hours)

    sorted_clients = sorted(clients, key=difficulty)

    # ── Pass 1: Main scheduling ──
    for client in sorted_clients:
        for day in client.days:
            if day not in client.schedule:
                continue

            need = client.schedule[day]
            remaining_start = need.start
            remaining_end = need.end
            block_num = 1
            pass1_subgaps = []

            attempts = 0
            while time_to_minutes(remaining_start) < time_to_minutes(remaining_end):
                attempts += 1
                if attempts > 30:
                    break

                result = try_assign_slot(
                    client, day, remaining_start, remaining_end,
                    therapists, timelines, assignments
                )
                if not result:
                    result = try_assign_slot(
                        client, day, remaining_start, remaining_end,
                        therapists, timelines, assignments, relaxed=True
                    )
                if result:
                    if time_to_minutes(result.start) > time_to_minutes(remaining_start) + 5:
                        pass1_subgaps.append((remaining_start, result.start))
                    result.notes = f"Block {block_num}"
                    assignments.append(result)
                    remaining_start = result.end
                    if time_to_minutes(remaining_start) < time_to_minutes(remaining_end):
                        block_num += 1
                else:
                    break

            # Fill pass 1 sub-gaps
            for sg_start, sg_end in pass1_subgaps:
                sg_remaining = sg_start
                sg_attempts = 0
                while time_to_minutes(sg_remaining) < time_to_minutes(sg_end):
                    sg_attempts += 1
                    if sg_attempts > 5:
                        break
                    result = try_assign_slot(
                        client, day, sg_remaining, sg_end,
                        therapists, timelines, assignments, relaxed=True
                    )
                    if result:
                        result.notes = f"Block {block_num} (sub-gap)"
                        assignments.append(result)
                        sg_remaining = result.end
                        block_num += 1
                    else:
                        break

    # ── Pass 2: Fill remaining gaps ──
    for client in sorted_clients:
        for day in client.days:
            if day not in client.schedule:
                continue

            need = client.schedule[day]
            # Find gaps in coverage
            day_a = sorted(
                [a for a in assignments if a.client == client.name and a.day == day],
                key=lambda a: a.start
            )

            gaps = []
            cursor = time_to_minutes(need.start)
            for a in day_a:
                a_start = time_to_minutes(a.start)
                if a_start > cursor:
                    gaps.append((minutes_to_time(cursor), a.start))
                cursor = max(cursor, time_to_minutes(a.end))
            if cursor < time_to_minutes(need.end):
                gaps.append((minutes_to_time(cursor), need.end))

            for gap_start, gap_end in gaps:
                remaining_start = gap_start
                block_num = len(day_a) + 1
                sub_gaps = []
                attempts = 0
                while time_to_minutes(remaining_start) < time_to_minutes(gap_end):
                    attempts += 1
                    if attempts > 15:
                        break
                    result = try_assign_slot(
                        client, day, remaining_start, gap_end,
                        therapists, timelines, assignments, relaxed=True
                    )
                    if result:
                        # Check if result starts after remaining_start (sub-gap)
                        if time_to_minutes(result.start) > time_to_minutes(remaining_start) + 5:
                            sub_gaps.append((remaining_start, result.start))
                        result.notes = f"Block {block_num} (gap fill)"
                        assignments.append(result)
                        remaining_start = result.end
                        block_num += 1
                    else:
                        break

                # Fill sub-gaps
                for sg_start, sg_end in sub_gaps:
                    sg_remaining = sg_start
                    sg_attempts = 0
                    while time_to_minutes(sg_remaining) < time_to_minutes(sg_end):
                        sg_attempts += 1
                        if sg_attempts > 5:
                            break
                        result = try_assign_slot(
                            client, day, sg_remaining, sg_end,
                            therapists, timelines, assignments, relaxed=True
                        )
                        if result:
                            result.notes = f"Block {block_num} (sub-gap fill)"
                            assignments.append(result)
                            sg_remaining = result.end
                            block_num += 1
                        else:
                            break

    # Collect final warnings for remaining gaps
    for client in sorted_clients:
        for day in client.days:
            if day not in client.schedule:
                continue
            need = client.schedule[day]
            day_a = sorted(
                [a for a in assignments if a.client == client.name and a.day == day],
                key=lambda a: a.start
            )
            cursor = time_to_minutes(need.start)
            for a in day_a:
                a_start = time_to_minutes(a.start)
                if a_start > cursor + 5:  # >5 min gap
                    warnings.append(
                        f"COVERAGE GAP: {client.name} on {day} "
                        f"{fmt_time(minutes_to_time(cursor))}-{fmt_time(a.start)}"
                    )
                cursor = max(cursor, time_to_minutes(a.end))
            if cursor < time_to_minutes(need.end) - 5:
                warnings.append(
                    f"COVERAGE GAP: {client.name} on {day} "
                    f"{fmt_time(minutes_to_time(cursor))}-{fmt_time(need.end)}"
                )

    return assignments, warnings


# ── Validation ───────────────────────────────────────────────────────────────

def validate_schedule(therapists: list, clients: list, assignments: list) -> list:
    warnings = []

    # 1. Overlap check
    for t in therapists:
        for day in DAYS_ORDER:
            day_a = therapist_day_assignments(t.name, day, assignments)
            for i in range(len(day_a)):
                for j in range(i + 1, len(day_a)):
                    a1, a2 = day_a[i], day_a[j]
                    if TimeBlock(a1.start, a1.end).overlaps(TimeBlock(a2.start, a2.end)):
                        warnings.append(
                            f"OVERLAP: {t.name} on {day} — "
                            f"{a1.client}({fmt_time(a1.start)}-{fmt_time(a1.end)}) vs "
                            f"{a2.client}({fmt_time(a2.start)}-{fmt_time(a2.end)})"
                        )

    # 2. Chain check (4h max without 30-min break)
    for t in therapists:
        for day in DAYS_ORDER:
            day_a = therapist_day_assignments(t.name, day, assignments)
            if len(day_a) < 1:
                continue
            chain_start_m = time_to_minutes(day_a[0].start)
            chain_end_m = time_to_minutes(day_a[0].end)
            for i in range(1, len(day_a)):
                a_start = time_to_minutes(day_a[i].start)
                gap = a_start - chain_end_m
                if gap < BREAK_MINUTES:
                    chain_end_m = time_to_minutes(day_a[i].end)
                else:
                    chain_hrs = (chain_end_m - chain_start_m) / 60.0
                    if chain_hrs > MAX_CHAIN_HOURS + 0.01:
                        warnings.append(
                            f"CHAIN: {t.name} on {day} — {chain_hrs:.1f}h continuous "
                            f"({fmt_time(minutes_to_time(chain_start_m))}-{fmt_time(minutes_to_time(chain_end_m))})"
                        )
                    chain_start_m = time_to_minutes(day_a[i].start)
                    chain_end_m = time_to_minutes(day_a[i].end)
            # Final chain
            chain_hrs = (chain_end_m - chain_start_m) / 60.0
            if chain_hrs > MAX_CHAIN_HOURS + 0.01:
                warnings.append(
                    f"CHAIN: {t.name} on {day} — {chain_hrs:.1f}h continuous "
                    f"({fmt_time(minutes_to_time(chain_start_m))}-{fmt_time(minutes_to_time(chain_end_m))})"
                )

    # 3. Client coverage check
    for c in clients:
        for day in c.days:
            if day not in c.schedule:
                continue
            need = c.schedule[day]
            day_a = sorted(
                [a for a in assignments if a.client == c.name and a.day == day],
                key=lambda a: a.start
            )
            if not day_a:
                warnings.append(f"NO COVERAGE: {c.name} on {day} {fmt_time(need.start)}-{fmt_time(need.end)}")
                continue

            gaps = []
            cursor = time_to_minutes(need.start)
            for a in day_a:
                a_start = time_to_minutes(a.start)
                if a_start > cursor:
                    gaps.append(f"{fmt_time(minutes_to_time(cursor))}-{fmt_time(a.start)}")
                cursor = max(cursor, time_to_minutes(a.end))
            if cursor < time_to_minutes(need.end):
                gaps.append(f"{fmt_time(minutes_to_time(cursor))}-{fmt_time(need.end)}")
            if gaps:
                warnings.append(f"GAPS: {c.name} on {day} — {', '.join(gaps)}")

    # 4. Workload check
    for t in therapists:
        weekly = therapist_weekly_hours(t.name, assignments)
        if weekly < 0.01:
            continue
        if weekly >= FORTY_CAP and not t.forty_hour_eligible:
            warnings.append(f"CRITICAL: {t.name} at {weekly:.1f}h (NOT 40h eligible)")
        elif weekly >= FORTY_CAP:
            warnings.append(f"40H: {t.name} at {weekly:.1f}h — verify 2h mid-day break")
        elif weekly >= HARD_CAP:
            warnings.append(f"HIGH: {t.name} at {weekly:.1f}h (hard cap)")
        elif weekly >= SOFT_CAP:
            warnings.append(f"WARN: {t.name} at {weekly:.1f}h (soft cap)")
        if t.preferred_max_hours and weekly > t.preferred_max_hours:
            warnings.append(f"PREF MAX: {t.name} at {weekly:.1f}h (pref {t.preferred_max_hours}h)")

    return warnings


# ── Output ───────────────────────────────────────────────────────────────────

def format_days_list(days: list) -> str:
    if not days:
        return ""
    # Sort by DAYS_ORDER
    days = sorted(set(days), key=lambda d: DAYS_ORDER.index(d) if d in DAYS_ORDER else 99)
    if days == WEEKDAYS:
        return "Mon-Fri"
    if days == DAYS_ORDER[:6]:
        return "Mon-Sat"

    indices = [DAYS_ORDER.index(d) for d in days if d in DAYS_ORDER]
    if not indices:
        return ", ".join(days)
    indices.sort()
    if len(indices) > 2 and indices == list(range(indices[0], indices[-1] + 1)):
        return f"{DAYS_ORDER[indices[0]]}-{DAYS_ORDER[indices[-1]]}"

    return ", ".join(days)


def group_assignments(assignments: list) -> list:
    """Group identical assignments across days into single rows."""
    assignments.sort(key=lambda a: (
        a.client, a.therapist, time_to_minutes(a.start),
        DAYS_ORDER.index(a.day) if a.day in DAYS_ORDER else 99
    ))

    groups = []
    used = set()

    for i, a in enumerate(assignments):
        if i in used:
            continue
        matching_days = [a.day]
        used.add(i)

        for j in range(i + 1, len(assignments)):
            if j in used:
                continue
            b = assignments[j]
            if (b.client == a.client and b.therapist == a.therapist and
                    b.start == a.start and b.end == a.end and
                    b.location == a.location and b.assignment_type == a.assignment_type):
                matching_days.append(b.day)
                used.add(j)

        matching_days.sort(key=lambda d: DAYS_ORDER.index(d) if d in DAYS_ORDER else 99)

        groups.append({
            'client': a.client,
            'therapist': a.therapist,
            'days': format_days_list(matching_days),
            'start': a.start,
            'end': a.end,
            'location': a.location,
            'type': a.assignment_type,
            'notes': a.notes
        })

    # Final sort: by client name, then start time
    groups.sort(key=lambda g: (g['client'], time_to_minutes(g['start']), g['therapist']))
    return groups


def write_output(wb, assignments: list):
    """Write schedule to 'Current Assignment' sheet."""
    ws = wb['Current Assignment']

    # Clear existing data (keep header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    grouped = group_assignments(assignments)

    for i, g in enumerate(grouped, start=2):
        ws.cell(row=i, column=1, value=g['client'])
        ws.cell(row=i, column=2, value=g['therapist'])
        ws.cell(row=i, column=3, value=g['days'])
        ws.cell(row=i, column=4, value=g['start'])
        ws.cell(row=i, column=5, value=g['end'])
        ws.cell(row=i, column=6, value=g['location'])
        ws.cell(row=i, column=7, value=g['type'])
        ws.cell(row=i, column=8, value=g['notes'])

    # Format time columns
    for row in ws.iter_rows(min_row=2, max_row=len(grouped) + 1, min_col=4, max_col=5):
        for cell in row:
            if cell.value:
                cell.number_format = 'h:mm AM/PM'

    print(f"  Wrote {len(grouped)} rows to 'Current Assignment' sheet")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.chdir('/Users/anjalparikh/Desktop/rbt-scheduler')

    print("=" * 60)
    print("  RBT Therapy Scheduler")
    print("=" * 60)

    print("\n[1/5] Loading data...")
    wb = openpyxl.load_workbook('RBT + Client Schedule.xlsx')

    therapists = load_therapists(wb['Therapists'])
    print(f"  Loaded {len(therapists)} therapists")

    clients = load_clients(wb['Clients'])
    print(f"  Loaded {len(clients)} clients")

    # Print summary
    print("\n── Therapists ──")
    for t in therapists:
        total = sum(sum(b.duration_hours() for b in blocks) for blocks in t.availability.values())
        flags = []
        if t.is_float:
            flags.append("FLOAT/LEAD")
        if t.flexible_days:
            flags.append(f"{t.flexible_days} days flex")
        if t.preferred_max_hours:
            flags.append(f"pref max {t.preferred_max_hours}h")
        tag = f" [{', '.join(flags)}]" if flags else ""
        print(f"  {t.name:<25} {format_days_list(t.days):<10} {total:>5.1f}h/week{tag}")

    print("\n── Clients ──")
    for c in clients:
        total = sum(tb.duration_hours() for tb in c.schedule.values())
        print(f"  {c.name:<12} {format_days_list(c.days):<10} {total:>5.1f}h/week "
              f"{'HIGH' if c.intensity == 'High' else 'Low ':>4} {c.location}")

    print("\n[2/5] Building therapist timelines...")
    timelines = build_therapist_timelines(therapists)

    print("[3/5] Scheduling...")
    assignments, sched_warnings = schedule_all(therapists, clients, timelines)
    print(f"  Generated {len(assignments)} individual assignments")
    if sched_warnings:
        for w in sched_warnings:
            print(f"  ⚠ {w}")

    print("\n[4/5] Validating...")
    val_warnings = validate_schedule(therapists, clients, assignments)
    all_warnings = sched_warnings + val_warnings
    if val_warnings:
        print(f"  {len(val_warnings)} validation issues:")
        for w in val_warnings:
            print(f"  ⚠ {w}")
    else:
        print("  All validations passed!")

    print("\n[5/5] Writing to Excel...")
    write_output(wb, assignments)
    wb.save('RBT + Client Schedule.xlsx')
    print("\nDone! Schedule saved to 'RBT + Client Schedule.xlsx' → 'Current Assignment' sheet")

    # Summary stats
    print("\n── Summary ──")
    total_client_hours = sum(
        tb.duration_hours() for c in clients for tb in c.schedule.values()
    )
    total_assigned = sum(a.duration_hours() for a in assignments)
    print(f"  Total client hours needed: {total_client_hours:.1f}h")
    print(f"  Total hours assigned:      {total_assigned:.1f}h")
    print(f"  Coverage:                  {total_assigned / total_client_hours * 100:.1f}%")
    print(f"  Warnings:                  {len(all_warnings)}")


if __name__ == '__main__':
    main()
